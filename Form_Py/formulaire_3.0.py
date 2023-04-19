import sys
from PyQt5.QtWidgets import *
from PyQt5.QtGui import QIcon, QPixmap
from PyQt5.QtCore import Qt, QDateTime
from datetime import datetime
from openpyxl import Workbook, load_workbook
import re
import os.path

class Formulaire(QMainWindow):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle("Formulaire d'inscription")
        self.setWindowIcon(QIcon("icon.png"))

        # Maximise la fenêtre
        self.setWindowState(Qt.WindowMaximized)

        # Ajout d'un fond d'écran
        background = QLabel(self)
        pixmap = QPixmap("background.png")
        background.setPixmap(pixmap)
        background.resize(self.width(), self.height())

        # Vérifier si le fichier Excel existe
        if os.path.isfile("formulaire.xlsx"):
            # Charger le fichier Excel existant
            fichier_excel = load_workbook("formulaire.xlsx")
            # Sélectionner la feuille active
            feuille = fichier_excel.active
            cellule_A1 = feuille.cell(row=1, column=11)
            self.ligne = int(cellule_A1.value)
        else:
            # Afficher un message d'erreur si le fichier Excel n'existe pas
            print("Le fichier formulaire.xlsx n'existe pas.")
            # Création du fichier Excel
            fichier_excel = Workbook()
            feuille = fichier_excel.active
            feuille.title = "Formulaire d'inscription"
            feuille.cell(row=1, column=1).value = "Nom"
            feuille.cell(row=1, column=2).value = "Prénom"
            feuille.cell(row=1, column=3).value = "E-mail"
            feuille.cell(row=1, column=4).value = "Jour"
            feuille.cell(row=1, column=5).value = "Heure"
            self.ligne = 2

        # Ajout des widgets
        self.nom_label = QLabel("Nom :", self)
        self.nom_label.move(100, 100)
        self.champ_nom = QLineEdit(self)
        self.champ_nom.setGeometry(200, 100, 200, 30)

        self.prenom_label = QLabel("Prénom :", self)
        self.prenom_label.move(100, 150)
        self.champ_prenom = QLineEdit(self)
        self.champ_prenom.setGeometry(200, 150, 200, 30)

        self.email_label = QLabel("Email :", self)
        self.email_label.move(100, 200)
        self.champ_email = QLineEdit(self)
        self.champ_email.setGeometry(200, 200, 200, 30)

        self.bouton_enregistrer = QPushButton("Enregistrer", self)
        self.bouton_enregistrer.setGeometry(100, 300, 120, 40)
        self.bouton_enregistrer.clicked.connect(self.enregistrer)

        self.bouton_effacer = QPushButton("Effacer", self)
        self.bouton_effacer.setGeometry(280, 300, 120, 40)
        self.bouton_effacer.clicked.connect(self.vider_champs)

        self.erreur_label = QLabel("", self)
        self.erreur_label.move(100, 350)

    # Fonction pour enregistrer les
    def enregistrer(self):
        nom = self.champ_nom.text()
        prenom = self.champ_prenom.text()
        email = self.champ_email.text()

        # Vérifier que tous les champs ont été remplis
        if nom == "" or prenom == "" or email == "":
            self.erreur_label.setText("Veuillez remplir tous les champs.")
            return

        # Vérifier que l'e-mail est valide
        if not re.match(r"[^@]+@[^@]+\.[^@]+", email):
            self.erreur_label.setText("Veuillez entrer une adresse e-mail valide.")
            return

        # Récupérer la date et l'heure actuelles
        date = datetime.now().strftime("%d/%m/%Y")
        heure = datetime.now().strftime("%H:%M:%S")

        # Enregistrer les données dans le fichier Excel
        self.feuille.cell(row=self.ligne, column=1).value = nom
        self.feuille.cell(row=self.ligne, column=2).value = prenom
        self.feuille.cell(row=self.ligne, column=3).value = email
        self.feuille.cell(row=self.ligne, column=4).value = date
        self.feuille.cell(row=self.ligne, column=5).value = heure
        self.ligne += 1
        # Enregistrer le numéro de ligne pour la prochaine inscription
        self.feuille.cell(row=1, column=11).value = self.ligne

        # Enregistrer les modifications dans le fichier Excel
        self.fichier_excel.save("formulaire.xlsx")

        # Effacer les champs de texte
        self.vider_champs()

        # Afficher un message de confirmation
        self.erreur_label.setStyleSheet("font: 16pt Arial; color: green;")
        self.erreur_label.setText("Les données ont été enregistrées avec succès.")

    def vider_champs(self):
        self.champ_nom.clear()
        self.champ_prenom.clear()
        self.champ_email.clear()
        self.erreur_label.clear()

def main():
    pass

if __name__ == '__main__':
    main()