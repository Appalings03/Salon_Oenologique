from tkinter import *
import re
import os.path
from openpyxl import *
from datetime import datetime

# Création de la fenêtre
fenetre = Tk()
fenetre.title("Formulaire d'inscription")

# Vérifier si le fichier Excel existe
if os.path.isfile("formulaire.xlsx"):
    # Charger le fichier Excel existant
    fichier_excel = load_workbook("formulaire.xlsx")
    # Sélectionner la feuille active
    feuille = fichier_excel.active
    cellule_A1 = feuille.cell(row=1, column=11)
    ligne = int(cellule_A1.value)
else:
    # Afficher un message d'erreur si le fichier Excel n'existe pas
    print("Le fichier formulaire.xlsx n'existe pas.")
    # Création du fichier Excel
    fichier_excel = Workbook()
    feuille = fichier_excel.active
    feuille.title = "Formulaire d'inscription"
    feuille.cell(row=1, column=1).value = "Nom"
    feuille.cell(row=1, column=2).value = "Prénom"
    feuille.cell(row=1, column=3).value = "E-mail"
    feuille.cell(row=1, column=4).value = "Jour"
    feuille.cell(row=1, column=5).value = "Heure"
    ligne = 2

# Fonction pour enregistrer les données du formulaire
def enregistrer():
    global ligne
    nom = champ_nom.get()
    prenom = champ_prenom.get()
    email = champ_email.get()

    if not nom or not prenom or not email:
        erreur_label.config(text="Veuillez remplir tous les champs.", fg="red")
    elif not re.match(r"[^@]+@[^@]+\.[^@]+", email):
        erreur_label.config(text="Adresse e-mail invalide.", fg="red")
    else:
        erreur_label.config(text="")
        # Enregistrer les données dans le fichier Excel
        feuille.cell(row=ligne, column=1).value = nom
        feuille.cell(row=ligne, column=2).value = prenom
        feuille.cell(row=ligne, column=3).value = email
        feuille.cell(row=ligne, column=4).value = datetime.now().strftime("%Y-%m-%d")
        feuille.cell(row=ligne, column=5).value = datetime.now().strftime("%H:%M:%S")
        ligne += 1
        feuille.cell(row=1, column=10).value = "Incrémentation"
        feuille.cell(row=1, column=11).value = ligne
        # Enregistrer le fichier Excel
        fichier_excel.save("formulaire.xlsx")
        print("Nom : ", nom)
        print("Prénom : ", prenom)
        print("Email : ", email)

        vider_champs()

def vider_champs():
    champ_nom.delete(0, END)
    champ_prenom.delete(0, END)
    champ_email.delete(0, END)

# Création des champs de formulaire
Label(fenetre, text="Nom : ").grid(row=0, column=0)
champ_nom = Entry(fenetre, width=30, font=("Arial", 16))
champ_nom.grid(row=0, column=1)

Label(fenetre, text="Prénom : ").grid(row=1, column=0)
champ_prenom = Entry(fenetre, width=30, font=("Arial", 16))
champ_prenom.grid(row=1, column=1)

Label(fenetre, text="Email : ").grid(row=2, column=0)
champ_email = Entry(fenetre, width=30, font=("Arial", 16))
champ_email.grid(row=2, column=1)

# Bouton pour enregistrer les données
bouton_enregistrer = Button(fenetre, text="Enregistrer", command=enregistrer)
bouton_enregistrer.grid(row=3, column=1)

# Bouton pour effacer les champs de texte
bouton_effacer = Button(fenetre, text="Effacer", command=vider_champs)
bouton_effacer.grid(row=3, column=2)

# Label pour afficher les erreurs
erreur_label = Label(fenetre, text="", fg="red", font=("Arial", 16))
erreur_label.grid(row=4, columnspan=2)

# Lancement de la fenêtre
fenetre.mainloop()