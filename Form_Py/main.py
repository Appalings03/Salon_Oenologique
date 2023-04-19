import tkinter as tk
from tkinter import *
import re
import os.path
from openpyxl import *
from datetime import datetime
import time

# Création de la fenêtre
fenetre = tk.Tk()
fenetre.title("Formulaire d'inscription")
# Maximise la fenêtre
fenetre.attributes("-fullscreen", True)


# Vérifier si le fichier Excel existe
if os.path.isfile("formulaire.xlsx"):
    # Charger le fichier Excel existant
    fichier_excel = load_workbook("formulaire.xlsx")
    # Sélectionner la feuille active
    feuille = fichier_excel.active
    cellule_A1 = feuille.cell(row=1, column=11)
    ligne = int(cellule_A1.value)
else:
    # Afficher un message d'erreur si le fichier Excel n'existe pas
    print("Le fichier formulaire.xlsx n'existe pas.")
    # Création du fichier Excel
    fichier_excel = Workbook()
    feuille = fichier_excel.active
    feuille.title = "Formulaire d'inscription"
    feuille.cell(row=1, column=1).value = "Nom"
    feuille.cell(row=1, column=2).value = "Prénom"
    feuille.cell(row=1, column=3).value = "E-mail"
    feuille.cell(row=1, column=4).value = "Jour"
    feuille.cell(row=1, column=5).value = "Heure"
    feuille.cell(row=1, column=6).value = "J'accepte"
    ligne = 2

# Fonction pour enregistrer les données du formulaire
def enregistrer():
    global ligne
    nom = champ_nom.get()
    prenom = champ_prenom.get()
    email = champ_email.get()
    if not nom or not prenom or not email:
        erreur_label.config(text="Veuillez remplir tous les champs.", fg="red")
    elif not re.match(r"[^@]+@[^@]+\.[^@]+", email):
        erreur_label.config(text="Adresse e-mail invalide.", fg="red")
    else:
        erreur_label.config(text="")
        # Enregistrer les données dans le fichier Excel
        feuille.cell(row=ligne, column=1).value = nom
        feuille.cell(row=ligne, column=2).value = prenom
        feuille.cell(row=ligne, column=3).value = email
        feuille.cell(row=ligne, column=4).value = datetime.now().strftime("%Y-%m-%d")
        feuille.cell(row=ligne, column=5).value = datetime.now().strftime("%H:%M:%S")
        feuille.cell(row=ligne, column=6).value = acceptation.get()
        ligne += 1
        feuille.cell(row=1, column=10).value = "Incrémentation"
        feuille.cell(row=1, column=11).value = ligne
        # Enregistrer le fichier Excel
        fichier_excel.save("formulaire.xlsx")
        print("Nom : ", nom)
        print("Prénom : ", prenom)
        print("Email : ", email)

        vider_champs()

def vider_champs():
    champ_nom.delete(0, tk.END)
    champ_prenom.delete(0, tk.END)
    champ_email.delete(0, tk.END)
    acceptation.set(False)

def onExit():
    exit_window = tk.Toplevel(fenetre)
    exit_window.title("Exit")

    password_entry = tk.Entry(exit_window, show="*")
    password_entry.pack()
    # Ajouter un widget Label pour afficher le résultat de la vérification du mot de passe
    global label
    label = tk.Label(fenetre, text="")
    label.pack()
    # Ajouter un bouton pour valider le mot de passe
    button_exit = tk.Button(exit_window, text="Valider", command=lambda: check_password(password_entry.get()))
    button_exit.pack()

# Fonction pour vérifier le mot de passe
def check_password(password):
    if password == "ventdange":
        label.config(text="Mot de passe correct!")
        time.sleep(2)
        fenetre.quit()
    else:
        label.config(text="Mot de passe incorrect!")

def update_button_state():
    if acceptation.get():
        bouton_enregistrer.config(state="normal")
    else:
        bouton_enregistrer.config(state="disabled")

toolbar = tk.Frame(fenetre, bg='grey', height=10)
toolbar.pack(side=tk.TOP, fill='x', anchor='n')

exit = tk.Button(toolbar, text='Exit', command=onExit)
exit.pack(side=tk.LEFT, padx=10)

# Création des champs de formulaire
nom_label = tk.Label(fenetre, text="Nom : ", font=("Arial", 16))
nom_label.pack(side=tk.TOP, pady=10)
champ_nom = tk.Entry(fenetre, width=30, font=("Arial", 16))
champ_nom.pack(side=tk.TOP, pady=10)

prenom_label = tk.Label(fenetre, text="Prénom : ", font=("Arial", 16))
prenom_label.pack(side=tk.TOP, pady=10)
champ_prenom = tk.Entry(fenetre, width=30, font=("Arial", 16))
champ_prenom.pack(side=tk.TOP, pady=10)

email_label = tk.Label(fenetre, text="Email : ", font=("Arial", 16))
email_label.pack(side=tk.TOP, pady=10)
champ_email = tk.Entry(fenetre, width=30, font=("Arial", 16))
champ_email.pack(side=tk.TOP, pady=10)

rgpd_text = """L’asbl les écrits vins tournent ces pages traite vos données dans le seul et unique but de vous fournir 
                    des informations concernant les activités de l’asbl.
        L’asbl s’engage à respecter la norme GDPR (global data protection regulation).
En cochant cette case vous acceptez la politique de confidentialité du traitement de vos données."""
rgpd_label = tk.Label(fenetre, text=rgpd_text,font=("Arial",12))
rgpd_label.pack(side=tk.TOP,pady=10)

# create a variable to store the checkbox value
acceptation = tk.BooleanVar()
# create the checkbox widget
checkbox = tk.Checkbutton(fenetre, text='J\'accepte les conditions', variable=acceptation, command=update_button_state, font=("Arial",16))
checkbox.pack(side=tk.TOP,pady=10)

# Bouton pour enregistrer les données
bouton_enregistrer = tk.Button(fenetre, text="Enregistrer", command=enregistrer, font=("Arial", 16),state="disabled")
bouton_enregistrer.pack(side=tk.LEFT, padx=50, pady=50)

# Bouton pour effacer les champs de texte
bouton_effacer = tk.Button(fenetre, text="Effacer", command=vider_champs, font=("Arial", 16))
bouton_effacer.pack(side=tk.RIGHT, padx=50, pady=50)

# Label pour afficher les erreurs
erreur_label = tk.Label(fenetre, text="", fg="red", font=("Arial", 16))
erreur_label.pack(side=tk.BOTTOM, pady=50)

# Lancement de la fenêtre
fenetre.mainloop()