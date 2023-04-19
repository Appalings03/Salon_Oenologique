import re
import os.path
from openpyxl import *
from datetime import datetime
from PyQt5.QtWidgets import *

class MainWindow(QWidget):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Formulaire d'inscription")

        # Maximise la fenêtre
        self.showMaximized()

        self.bg_image = QLabel(self)
        self.setStyleSheet("background-image:url(background.png);")
        self.bg_image.setScaledContents(True)

        # Vérifier si le fichier Excel existe
        if os.path.isfile("formulaire.xlsx"):
            # Charger le fichier Excel existant
            self.fichier_excel = load_workbook("formulaire.xlsx")
            # Sélectionner la feuille active
            self.feuille = self.fichier_excel.active
            cellule_A1 = self.feuille.cell(row=1, column=11)
            self.ligne = int(cellule_A1.value)
        else:
            # Afficher un message d'erreur si le fichier Excel n'existe pas
            print("Le fichier formulaire.xlsx n'existe pas.")
            # Création du fichier Excel
            self.fichier_excel = Workbook()
            self.feuille = self.fichier_excel.active
            self.feuille.title = "Formulaire d'inscription"
            self.feuille.cell(row=1, column=1).value = "Nom"
            self.feuille.cell(row=1, column=2).value = "Prénom"
            self.feuille.cell(row=1, column=3).value = "E-mail"
            self.feuille.cell(row=1, column=4).value = "Jour"
            self.feuille.cell(row=1, column=5).value = "Heure"
            self.ligne = 2

        # Création des éléments d'interface graphique
        self.setWindowTitle("Formulaire d'inscription")
        self.setFixedSize(800, 600)
        self.setStyleSheet("background-image: url(background.png);")

        # Création des champs de formulaire
        self.nom_label = QLabel("Nom : ")
        self.nom_label.setStyleSheet("font: 16pt Arial;")
        self.champ_nom = QLineEdit()
        self.champ_nom.setStyleSheet("font: 16pt Arial;")
        self.prenom_label = QLabel("Prénom : ")
        self.prenom_label.setStyleSheet("font: 16pt Arial;")
        self.champ_prenom = QLineEdit()
        self.champ_prenom.setStyleSheet("font: 16pt Arial;")
        self.email_label = QLabel("E-mail : ")
        self.email_label.setStyleSheet("font: 16pt Arial;")
        self.champ_email = QLineEdit()
        self.champ_email.setStyleSheet("font: 16pt Arial;")

        # Bouton pour enregistrer les données
        self.bouton_enregistrer = QPushButton("Enregistrer")
        self.bouton_enregistrer.setStyleSheet("font: 16pt Arial;")
        self.bouton_enregistrer.clicked.connect(self.enregistrer)

        # Bouton pour effacer les champs de texte
        self.bouton_effacer = QPushButton("Effacer")
        self.bouton_effacer.setStyleSheet("font: 16pt Arial;")
        self.bouton_effacer.clicked.connect(self.vider_champs)

        # Label pour afficher les erreurs
        self.erreur_label = QLabel("")
        self.erreur_label.setStyleSheet("font: 16pt Arial; color: red;")

        # Mise en page
        self.layout_horizontal_nom = QHBoxLayout()
        self.layout_horizontal_nom.addWidget(self.nom_label)
        self.layout_horizontal_nom.addWidget(self.champ_nom)
        self.layout_horizontal_prenom = QHBoxLayout()
        self.layout_horizontal_prenom.addWidget(self.prenom_label)
        self.layout_horizontal_prenom.addWidget(self.champ_prenom)

        self.layout_horizontal_email = QHBoxLayout()
        self.layout_horizontal_email.addWidget(self.email_label)
        self.layout_horizontal_email.addWidget(self.champ_email)

        self.layout_horizontal_boutons = QHBoxLayout()
        self.layout_horizontal_boutons.addWidget(self.bouton_enregistrer)
        self.layout_horizontal_boutons.addWidget(self.bouton_effacer)

        self.layout_vertical = QVBoxLayout()
        self.layout_vertical.addStretch()
        self.layout_vertical.addLayout(self.layout_horizontal_nom)
        self.layout_vertical.addLayout(self.layout_horizontal_prenom)
        self.layout_vertical.addLayout(self.layout_horizontal_email)
        self.layout_vertical.addWidget(self.erreur_label)
        self.layout_vertical.addStretch()
        self.layout_vertical.addLayout(self.layout_horizontal_boutons)

        self.setLayout(self.layout_vertical)

    def enregistrer(self):
        nom = self.champ_nom.text()
        prenom = self.champ_prenom.text()
        email = self.champ_email.text()

        # Vérifier que tous les champs ont été remplis
        if nom == "" or prenom == "" or email == "":
            self.erreur_label.setText("Veuillez remplir tous les champs.")
            return

        # Vérifier que l'e-mail est valide
        if not re.match(r"[^@]+@[^@]+\.[^@]+", email):
            self.erreur_label.setText("Veuillez entrer une adresse e-mail valide.")
            return

        # Récupérer la date et l'heure actuelles
        date = datetime.now().strftime("%d/%m/%Y")
        heure = datetime.now().strftime("%H:%M:%S")

        # Enregistrer les données dans le fichier Excel
        self.feuille.cell(row=self.ligne, column=1).value = nom
        self.feuille.cell(row=self.ligne, column=2).value = prenom
        self.feuille.cell(row=self.ligne, column=3).value = email
        self.feuille.cell(row=self.ligne, column=4).value = date
        self.feuille.cell(row=self.ligne, column=5).value = heure
        self.ligne += 1
        # Enregistrer le numéro de ligne pour la prochaine inscription
        self.feuille.cell(row=1, column=11).value = self.ligne

        # Enregistrer les modifications dans le fichier Excel
        self.fichier_excel.save("formulaire.xlsx")

        # Effacer les champs de texte
        self.vider_champs()

        # Afficher un message de confirmation
        self.erreur_label.setStyleSheet("font: 16pt Arial; color: green;")
        self.erreur_label.setText("Les données ont été enregistrées avec succès.")

    def vider_champs(self):
        self.champ_nom.clear()
        self.champ_prenom.clear()
        self.champ_email.clear()
        self.erreur_label.clear()
        