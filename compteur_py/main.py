import sys
import openpyxl
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QLabel, QVBoxLayout, QHBoxLayout
from datetime import datetime

class MyWidget(QWidget):
    def __init__(self, initial_counters):
        super().__init__()
        self.counters = initial_counters
        self.initUI()

    def initUI(self):
        # Définir les noms des labels pour chaque compteur
        self.counter_names = ['Payant', 'Francois', 'Vignerons', 'Alsec', 'Organisateurs']

        # Créer les boutons et les labels pour chaque compteur
        self.button_label_layouts = []
        for i, counter in enumerate(self.counters):
            button = QPushButton()
            label = QLabel(f'{self.counter_names[i]} : {counter}')

            # Configurer le texte du bouton en fonction de l'index
            if i == 0:
                button.setText('Payant')
            elif i == 1:
                button.setText('Francois')
            elif i == 2:
                button.setText('Vigneron')
            elif i == 3:
                button.setText('Alsec')
            elif i == 4:
                button.setText('Organisateurs')

            # Connecter le bouton à une fonction de rappel
            button.clicked.connect(lambda _, index=i: self.increment_counter(index))

            # Créer un layout horizontal pour le bouton et le label et ajouter le bouton et le label au layout
            button_label_layout = QHBoxLayout()
            button_label_layout.addWidget(button)
            button_label_layout.addWidget(label)

            # Ajouter le layout à la liste des layouts de bouton et de label
            self.button_label_layouts.append(button_label_layout)

        # Créer un layout vertical et ajouter les layouts de bouton et de label, ainsi que le label somme
        vbox = QVBoxLayout()
        for button_label_layout in self.button_label_layouts:
            vbox.addLayout(button_label_layout)
        self.sum_label = QLabel('Total : ' + str(sum(self.counters)))
        vbox.addWidget(self.sum_label)

        # Appliquer le layout à la fenêtre
        self.setLayout(vbox)

    def increment_counter(self, index):
        # Incrémenter le compteur correspondant et mettre à jour le label correspondant
        self.counters[index] += 1
        self.button_label_layouts[index].itemAt(1).widget().setText(f'{self.counter_names[index]} : {self.counters[index]}')

        # Enregistrer la nouvelle valeur du compteur et les informations sur le bouton pressé dans la feuille de calcul "Logs"
        workbook = openpyxl.load_workbook('relevé_entrée.xlsx')
        worksheet = workbook.active
        for i, counter in enumerate(self.counters):
            worksheet.cell(row=i+1, column=1, value=self.button_label_layouts[i].itemAt(0).widget().text())
            worksheet.cell(row=i+1, column=2, value=str(counter))

        # Ajouter une nouvelle ligne à la feuille de calcul "Logs" avec le nom du bouton, la date et l'heure actuelles
        logs_worksheet = workbook['Logs']
        log_row = [self.counter_names[index], datetime.now().strftime('%Y-%m-%d'), datetime.now().strftime('%H:%M:%S')]
        logs_worksheet.append(log_row)

        workbook.save('relevé_entrée.xlsx')
        workbook.close()

        # Mettre à jour le label somme
        self.sum_label.setText('Total : ' + str(sum(self.counters)))

if __name__ == '__main__':
    app = QApplication(sys.argv)

    # Vérifier si le fichier Excel existe et récupérer les valeurs initiales des compteurs
    try:
        workbook = openpyxl.load_workbook('relevé_entrée.xlsx')
        worksheet = workbook.active
        compteur = workbook['Compte']
        initial_counters = [int(compteur.cell(row=i+1, column=2).value) for i in range(compteur.max_row)]
        workbook.close()

    except FileNotFoundError:
        # Si le fichier n'existe pas, initialiser les compteurs à zéro et créer le fichier Excel
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Compte'
        for i in range(5):
            button_name = ['Payant', 'Francois', 'Vignerons', 'Alsec','Organisateurs'][i]
            worksheet.cell(row=i+1, column=1, value=button_name)
            worksheet.cell(row=i+1, column=2, value='0')
        
        # Add a second worksheet with columns for Nom, Date, and Heure
        worksheet_logs = workbook.create_sheet(title='Logs')
        worksheet_logs.cell(row=1, column=1, value='Log')
        worksheet_logs.cell(row=1, column=2, value='Date')
        worksheet_logs.cell(row=1, column=3, value='Heure')
        workbook.save('relevé_entrée.xlsx')
        workbook.close()
        initial_counters = [0, 0, 0, 0, 0]

    widget = MyWidget(initial_counters)
    widget.show()
    sys.exit(app.exec_())