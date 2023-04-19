import sys
import openpyxl
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QLabel, QVBoxLayout, QHBoxLayout
from datetime import datetime

class MyWidget(QWidget):
    def __init__(self, initial_counters):
        super().__init__()
        self.counters = initial_counters
        self.initUI()

    def initUI(self):
        # Définir les noms des labels pour chaque compteur
        self.counter_names = ['Compteur Payant', 'Compteur Inscription Francois', 'Compteur Inscription Vigneron', 'Compteur Autre']

        # Créer les boutons et les labels pour chaque compteur
        self.button_label_layouts = []
        for i, counter in enumerate(self.counters):
            button = QPushButton()
            label = QLabel(f'{self.counter_names[i]} : {counter}')

            # Configurer le texte du bouton en fonction de l'index
            if i == 0:
                button.setText('Payant')
            elif i == 1:
                button.setText('Inscription Francois')
            elif i == 2:
                button.setText('Inscription Vigneron')
            elif i == 3:
                button.setText('Autre')

            # Connecter le bouton à une fonction de rappel
            button.clicked.connect(lambda _, index=i: self.increment_counter(index))

            # Créer un layout horizontal pour le bouton et le label et ajouter le bouton et le label au layout
            button_label_layout = QHBoxLayout()
            button_label_layout.addWidget(button)
            button_label_layout.addWidget(label)

            # Ajouter le layout à la liste des layouts de bouton et de label
            self.button_label_layouts.append(button_label_layout)

        # Créer un layout vertical et ajouter les layouts de bouton et de label, ainsi que le label somme
        vbox = QVBoxLayout()
        for button_label_layout in self.button_label_layouts:
            vbox.addLayout(button_label_layout)
        self.sum_label = QLabel('Total : ' + str(sum(self.counters)))
        vbox.addWidget(self.sum_label)

        # Appliquer le layout à la fenêtre
        self.setLayout(vbox)

    def increment_counter(self, index):
        # Incrémenter le compteur correspondant et mettre à jour le label correspondant
        self.counters[index] += 1
        self.button_label_layouts[index].itemAt(1).widget().setText(f'{self.counter_names[index]} : {self.counters[index]}')

        # Enregistrer la nouvelle valeur du compteur dans le fichier Excel
        workbook = openpyxl.load_workbook('compteurs.xlsx')
        worksheet = workbook.active
        for i, counter in enumerate(self.counters):
            worksheet.cell(row=i+1, column=1, value=self.button_label_layouts[i].itemAt(0).widget().text())
            worksheet.cell(row=i+1, column=2, value=str(counter))

        # Vérifier si la feuille "Boutons" existe déjà
        if "Evenements" in workbook.sheetnames:
            worksheet_boutons = workbook["Evenements"]
        else:
            worksheet_boutons = workbook.create_sheet("Evenements")

        # Ajouter une nouvelle ligne à la feuille "Boutons" avec le nom du bouton, la date et l'heure
        row_num = worksheet_boutons.max_row + 1
        button_name = ['Payant', 'Inscription Francois', 'Inscription Vigneron', 'Autre'][index]
        worksheet_boutons.cell(row=row_num, column=1, value=button_name)
        worksheet_boutons.cell(row=row_num, column=2, value=datetime.now().strftime("%Y-%m-%d"))
        worksheet_boutons.cell(row=row_num, column=3, value=datetime.now().strftime("%H:%M:%S"))

        workbook.save('compteurs.xlsx')
        workbook.close()

        # Mettre à jour le label somme
        self.sum_label.setText('Total : ' + str(sum(self.counters)))


if __name__ == '__main__':
    app = QApplication(sys.argv)

    # Vérifier si le fichier Excel existe et récupérer les valeurs initiales des compteurs
    try:
        workbook = openpyxl.load_workbook('compteurs.xlsx')
    except FileNotFoundError:
        # Si le fichier n'existe pas, initialiser les compteurs à zéro et créer le fichier Excel
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Compteurs'
        worksheet.cell(row=1, column=1, value='Nom')
        worksheet.cell(row=1, column=2, value='Compteur')
        worksheet2 = workbook.create_sheet('Evenements')
        worksheet2.cell(row=1, column=1, value='Nom')
        worksheet2.cell(row=1, column=2, value='Date')
        worksheet2.cell(row=1, column=3, value='Heure')
        for i in range(4):
            button_name = ['Payant', 'Inscription Francois', 'Inscription Vigneron', 'Autre'][i]
            worksheet.cell(row=i+2, column=1, value=button_name)
            worksheet.cell(row=i+2, column=2, value='0')
        workbook.save('compteurs.xlsx')

    worksheet = workbook.active
    worksheet2 = workbook['Evenements']
    initial_counters = [int(worksheet.cell(row=i+2, column=2).value) for i in range(4)]

    widget = MyWidget(initial_counters)
    widget.show()
    sys.exit(app.exec_())