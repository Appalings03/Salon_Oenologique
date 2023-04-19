from PySide6.QtWidgets import QMainWindow,QWidget, QToolBar, QLabel, QSizePolicy, QHBoxLayout, QVBoxLayout, QPushButton, QCheckBox, QLineEdit
from PySide6.QtCore import QSize, Qt
from PySide6.QtGui import QAction, QFont
import re
import os.path
from openpyxl import *
from datetime import datetime


class MainWindow(QMainWindow):
    def __init__(self,app,champ=['Nom', 'Prénom', 'E-Mail']):
        super().__init__()
        self.app = app
        self.champs = champ
        print(self.champs)
        self.initUI()
        if os.path.isfile("formulaire.xlsx"):
            # Charger le fichier Excel existant
            self.fichier_excel = load_workbook("formulaire.xlsx")
            # Sélectionner la feuille active
            self.feuille = self.fichier_excel.active
            cellule_A1 = self.feuille.cell(row=1, column=11)
            self.ligne = int(cellule_A1.value)
            self.fichier_excel.close()
        else:
            # Afficher un message d'erreur si le fichier Excel n'existe pas
            print("Le fichier formulaire.xlsx n'existe pas.")
            # Création du fichier Excel
            self.fichier_excel = Workbook()
            self.feuille = self.fichier_excel.active
            self.feuille.title = "Formulaire d'inscription"
            self.feuille.cell(row=1, column=1).value = "Nom"
            self.feuille.cell(row=1, column=2).value = "Prénom"
            self.feuille.cell(row=1, column=3).value = "E-mail"
            self.feuille.cell(row=1, column=4).value = "Jour"
            self.feuille.cell(row=1, column=5).value = "Heure"
            self.ligne = 2
            self.fichier_excel.close()

    def initUI(self):
        self.setWindowTitle("Formulaire Salon Oenologique")
        self.setup_toolbar()
        self.setup_button()
    
    def setup_button(self):
        vbox = QVBoxLayout(alignment=Qt.AlignCenter)
        for champ in self.champs:
            hbox = QHBoxLayout(alignment=Qt.AlignCenter)
            label = QLabel(champ)
            label.setFont(QFont("Arial", 16)) 
            line_edit = QLineEdit()
            line_edit.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
            line_edit.setFont(QFont("Arial", 16))
            line_edit.textChanged.connect(lambda text, key=champ: self.get_text(text, key)) # connect to get_text function
            
            hbox.addWidget(label)
            hbox.addWidget(line_edit)
            vbox.addLayout(hbox)
        #Error Label
        error_label = QLabel(alignment=Qt.AlignCenter)
        error_label.setFont(QFont("Arial", 16))
        vbox.addWidget(error_label)
        #RGPD
        policy_hbox = QHBoxLayout(alignment=Qt.AlignCenter)
        rgpd_text = """L’asbl les écrits vins tournent cepagestraite vos données dans le seul et unique but de vous fournir 
                    des informations concernant les activités de l’asbl.
        L’asbl s’engage à respecter la norme GDPR (global data protection regulation).
En cochant cette case vous acceptez la politique de confidentialité du traitement de vos données."""

        rgpd_label = QLabel(rgpd_text)

        rgpd_label.setFont(QFont("Arial", 12))
        rgpd_checkbox = QCheckBox()
        rgpd_checkbox.setStyleSheet("QCheckBox::indicator { width: 20px; height: 20px; }")
        policy_hbox.addWidget(rgpd_label)
        policy_hbox.addWidget(rgpd_checkbox)
        vbox.addLayout(policy_hbox)

        bouton_enregistrer = QPushButton("Enregistrer")
        bouton_enregistrer.clicked.connect(self.enregistrer)
        vbox.addWidget(bouton_enregistrer)
        
        widget = QWidget()
        widget.setLayout(vbox)
        self.setCentralWidget(widget)
        
    def setup_toolbar(self):
        #Toolbar création
        toolbar = QToolBar("Main Toolbar")
        toolbar.setIconSize(QSize(16,16))
        self.addToolBar(toolbar)
        #Quit action Création
        quit_action = toolbar.addAction('Quit')
        quit_action.triggered.connect(self.quit_app)
        toolbar.addAction(quit_action)
        #FullScreen Action Création
        full_screen_action = QAction("Full Screen", self)
        full_screen_action.setCheckable(True)
        full_screen_action.setChecked(False)
        full_screen_action.triggered.connect(self.toggle_full_screen)
        toolbar.addAction(full_screen_action)

    def toggle_full_screen(self, checked):
        if checked:
            self.showFullScreen()
        else:
            self.showNormal()

    def quit_app(self):
        self.app.quit()
    
    def enregistrer(self):
        nom = self.get_text(self.champs[0])
        prenom = self.get_text(self.champs[1])
        email = self.get_text(self.champs[2])

        if not nom or not prenom or not email:
            self.erreur_label.setText("Veuillez remplir tous les champs.")
            self.erreur_label.setStyleSheet("color: red")
        elif not re.match(r"[^@]+@[^@]+\.[^@]+", email):
            self.erreur_label.setText("Adresse e-mail invalide.")
            self.erreur_label.setStyleSheet("color: red")
        else:
            self.erreur_label.setText("")
            self.fichier_excel = load_workbook("formulaire.xlsx")
            self.feuille = self.fichier_excel.active
            # Enregistrer les données dans le fichier Excel
            self.feuille.cell(row=self.ligne, column=1).value = nom
            self.feuille.cell(row=self.ligne, column=2).value = prenom
            self.feuille.cell(row=self.ligne, column=3).value = email
            self.feuille.cell(row=self.ligne, column=4).value = datetime.now().strftime("%Y-%m-%d")
            self.feuille.cell(row=self.ligne, column=5).value = datetime.now().strftime("%H:%M:%S")
            self.ligne += 1
            self.feuille.cell(row=1, column=10).value = "Incrémentation"
            self.feuille.cell(row=1, column=11).value = self.ligne
            # Enregistrer le fichier Excel
            self.fichier_excel.save("formulaire.xlsx")
            print("Nom : ", nom)
            print("Prénom : ", prenom)
            print("Email : ", email)
            self.fichier_excel.close()

            self.vider_champs()

    def viderChamps(self):
        self.champ_nom.setText("")
        self.champ_prenom.setText("")
        self.champ_email.setText("")

    def get_text(self, text, key):
        self.textes[key] = text # store text in dictionary with the corresponding key
        print(self.textes)